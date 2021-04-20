import openpyxl
import re

def find_loc(row, info, info_loc):
    for val in info:
        i = 0
        for cell in row:
            i += 1
            if cell.value == val:
                info_loc.append(i)
                break
    return info_loc

# def compare_info(compare_check, compare_annual):
#     if len(compare_check) != len(compare_check):
#         print("长度不同，比较失败")
#         return 2
#     compare_check_0 = str(compare_check[0])
#     compare_check_1 = str(compare_check[1])
#     compare_annual_0 = str(compare_annual[0])
#     compare_annual_1 = str(compare_annual[1])
#     print("strrr:\n", "compare_check_0:", compare_check_0, "compare_check_1:", compare_check_1, "\ncompare_annual_0", compare_annual_0,
#         "compare_annual_1", compare_annual_1)
#     if bool(re.search(r'\d', compare_check_0)) == True:
#         pattern = re.compile(u'[\u4e00-\u9fa5]')
#         compare_check_0 = re.sub(pattern, '', compare_check_0)
#         compare_check_0 = re.sub('\\（.*?\\）', '', compare_check_0)

#     compare_check_0 = compare_check_0.replace(' ', '')
#     compare_check_1 = compare_check_1.replace(' ', '')
#     compare_annual_0 = compare_annual_0.replace(' ', '')
#     compare_annual_1 = compare_annual_1.replace(' ', '')

#     if (compare_check_0 == compare_annual_0) or (compare_check_0 in compare_annual_0) or (compare_annual_0 in compare_check_0):
#         if (compare_check_1 in compare_annual_1) or (compare_annual_1 in compare_check_1):
#             print("匹配成功1:", "compare_check_0:", compare_check_0, "compare_check_1:", compare_check_1, "\ncompare_annual_0", compare_annual_0,
#                 "compare_annual_1", compare_annual_1)
#             print(compare_check, compare_annual)
#             return 1
#         elif (compare_check_1 == "其它") and (compare_annual_1 in compare_check[0]):
#             print("匹配成功2")
#             print(compare_check, compare_annual)
#             return 1
#         else:
#             pass
#     else:
#         pass

#     return 0

def compare_info(compare_check, compare_annual):
    if len(compare_check) != len(compare_check):
        print("长度不同，比较失败")
        return 2
    compare_check_0 = str(compare_check[0])
    compare_check_1 = str(compare_check[1])
    compare_annual_0 = str(compare_annual[0])
    compare_annual_1 = str(compare_annual[1])

    if (compare_check_0 in compare_annual_0) or (compare_annual_0 in compare_check_0):
        if (compare_check_1 in compare_annual_1) or (compare_annual_1 in compare_check_1):
            print("匹配成功1")
            print(compare_check, compare_annual)
            return 1
        elif compare_annual_1 in compare_check[0]:
            print("匹配成功2")
            print(compare_check, compare_annual)
            return 1
        else:
            pass
    else:
        pass

    return 0

def deal_success(sheet_ch, sheet_an, row_ch, row_an):
    print("\n")
    print("赋值*********")
    print("row_ch:", row_ch, "\trow_an", row_an)

    cell_an = sheet_an.cell(row=int(row_an), column=14).value
    print("待赋值：",cell_an)

    sheet_ch.cell(row=row_ch, column=8, value=cell_an)
    print("赋值:", sheet_ch.cell(row=row_ch, column=8).value)
    print("\n")
    return


def take_info(sheet_ch, sheet_an):
    rowmax_ch = sheet_ch.max_row
    rowmax_an = sheet_an.max_row
    com_check = []
    com_annual = []
    success_num = 0
    for item_ch in range(5, rowmax_ch+1):
        del com_check[-2:]
        for i in range(0, len(check_info_loc)):
                com_check.append(sheet_ch.cell(row=item_ch, column=check_info_loc[i]).value)
        for item_an in range(4, rowmax_an+1):
            del com_annual[-2:]
            for i in range(0, len(annual_info_loc)):
                com_annual.append(sheet_an.cell(row=item_an, column=annual_info_loc[i]).value)

            # print("检测提取是否正确", com_check, com_annual)
            if compare_info(com_check, com_annual) == 1:
                success_num += 1
                print("匹配成功次数：", success_num)
                print("\nann第",item_an,"次循环")
                deal_success(sheet_ch, sheet_an, item_ch, item_an)
                break
            else:
                # print("匹配失败")
                continue
    return

if __name__ =="__main__":

    # 昆安
    check_info = ['桥名', '桥幅']
    check_info_loc = []

    annual_info = ['桥位中心桩号', '位置'] 
    annual_info_loc = []

    input_item = ["桥梁固定编码"]
    input_loc = [2]
    output_loc = 8

    book_check = openpyxl.load_workbook("content/检测数据-20210410/云路/桥梁动态数据-定期检查（云南云路工程检测有限公司-2020）.xlsx")
    book_annual = openpyxl.load_workbook("content/交投桥梁基础数据-养护定检与年报编码匹配明细表2021.04.12（全部数据）.xlsx")

    sheet_check = book_check.active
    sheet_annual = book_annual["昆西-昆安31"]

    print(sheet_check)
    print(sheet_annual)

    find_loc(sheet_check[3], check_info, check_info_loc)
    find_loc(sheet_annual[3], annual_info, annual_info_loc)
    find_loc(sheet_annual[3], input_item, input_loc)

    print(check_info_loc)
    print(annual_info_loc)
    print(input_loc)

    # print("rowmax_ch:", sheet_check.max_row, "\trowmax_an:", sheet_annual.max_row)


    take_info(sheet_check, sheet_annual)
    book_check.save(filename = '桥梁动态数据-定期检查（云南云路工程检测有限公司-2020）_3.xlsx') 

