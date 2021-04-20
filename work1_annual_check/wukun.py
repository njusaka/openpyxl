import openpyxl
import re

def find_loc(row, info, info_loc):
    for val in info:
        i = 0
        for cell in row:
            i += 1
            if cell.value == val:
                info_loc.append(i)
                break
    return info_loc

def compare_info(compare_check, compare_annual):
    if len(compare_check) != len(compare_check):
        print("长度不同，比较失败")
        return 2

    compare_check_0 = str(compare_check[0])
    compare_check_0 = compare_check_0.upper()
    compare_check_0_special = re.sub('\\(.*?\\)', '', compare_check_0)
    compare_check_0_special = re.sub('\\（.*?\\）', '', compare_check_0_special)
    compare_check_0_special = re.sub(u'[\u4e00-\u9fa5]', '', compare_check_0_special) #去中文

    compare_check_0_special_take = re.sub(u'[^\u4e00-\u9fa5]', '', compare_check_0)  #取中文

    compare_check_1 = str(compare_check[1])

    compare_annual_0 = str(compare_annual[0])
    pattern = re.compile(r'[（](.*?)[）]')
    temp = re.findall(pattern, compare_annual_0)
    # print("检验取括号里的值：temp", temp, '\t\t', compare_check, compare_annual)
    compare_annual_0_brack = temp[0]
    compare_annual_0_brack = re.sub(u'[\u4e00-\u9fa5]', '', compare_annual_0_brack) #去掉括号里的中文
    # print("compare_check_0_special", compare_check_0_special)
    compare_annual_0_special = re.sub('\\(.*?\\)', '', compare_annual_0)
    compare_annual_0_special = re.sub('\\（.*?\\）', '', compare_annual_0_special)

    compare_annual_1 = str(compare_annual[1])

    # print("test: ", compare_check_0, compare_check_1, compare_annual_0, compare_annual_1, compare_annual_0_brack)
    # print(compare_check_0_special, compare_annual_0_special)

    if (compare_check_0 in compare_annual_0) or (compare_annual_0 in compare_check_0) \
        or (compare_annual_0_brack in compare_check_0) or (compare_check_0_special == compare_annual_0_special)\
            or (compare_check_0_special in compare_annual_0):
        if compare_check_1 != '':
            if compare_check_1 in compare_annual_1 or compare_check_1 in compare_annual_1:
                # print("匹配成功1:", compare_check_0, compare_check_1, compare_annual_0, compare_annual_1)
                return 1
            #其它， 单幅
            elif compare_annual_1 in compare_check_0:
                # print("匹配成功2", compare_check_0, compare_check_1, compare_annual_0, compare_annual_1)
                return 1
            elif compare_check_0_special_take =='跨线桥' and compare_annual_1 == '上跨桥':
                # print("匹配成功上跨桥", compare_check_0, compare_check_1, compare_annual_0, compare_annual_1)
                return 1
            elif compare_annual_1 == '解家营立交匝道':
                # print("匹配成功解家营", compare_check_0, compare_check_1, compare_annual_0, compare_annual_1)
                return 1
            else:
                # print("匹配失败2:", compare_check_0, compare_check_1, compare_annual_0, compare_annual_1)
                pass
        elif compare_annual_1 in compare_check_0:
            # print("匹配成功3", compare_check_0, compare_check_1, compare_annual_0, compare_annual_1)
            return 1
        else:
            # print("匹配失败3:", compare_check_0, compare_check_1, compare_annual_0, compare_annual_1)
            pass
    else:
        # print("匹配失败1:", compare_check_0, compare_check_1, compare_annual_0, compare_annual_1)
        pass

    return 0

def deal_success(sheet_ch, sheet_an, row_ch, row_an):
    print("\n")
    print("赋值*********")
    print("row_ch:", row_ch, "\trow_an", row_an)

    cell_an = sheet_an.cell(row=int(row_an), column=16).value   #年报编号
    print("待赋值：",cell_an)

    sheet_ch.cell(row=row_ch, column=8, value=cell_an)
    print("赋值:", sheet_ch.cell(row=row_ch, column=8).value)
    print("\n")
    return


def take_info(sheet_ch, sheet_an):
    rowmax_ch = sheet_ch.max_row
    rowmax_an = sheet_an.max_row
    com_check = []
    com_annual = []
    success_num = 0
    for item_ch in range(5, rowmax_ch+1):            #range(5, rowmax_ch+1)
        del com_check[-2:]
        for i in range(0, len(check_info_loc)):
                com_check.append(sheet_ch.cell(row=item_ch, column=check_info_loc[i]).value)
        for item_an in range(4, rowmax_an):  #180行是空行  range(4, rowmax_an)
            del com_annual[-2:]
            for i in range(0, len(annual_info_loc)):
                com_annual.append(sheet_an.cell(row=item_an, column=annual_info_loc[i]).value)
            # print("内层第{}行".format(item_an))
            # print("检测提取是否正确", com_check, com_annual)
            if compare_info(com_check, com_annual) == 1:
                success_num += 1
                print("匹配成功次数：", success_num)
                deal_success(sheet_ch, sheet_an, item_ch, item_an)
                break
            else:
                # print("匹配失败")
                continue
    return

if __name__ =="__main__":

     # 武昆
    check_info = ['桥名', '桥幅']
    check_info_loc = []

    annual_info = ['国高网中心桩号', '位置'] 
    annual_info_loc = []

    input_item = ["年报编码"]
    input_loc = [16]
    output_loc = 8

    book_check = openpyxl.load_workbook("content/桥梁动态数据-定期检查（云南云路工程检测有限公司-2020）_f5.xlsx")
    book_annual = openpyxl.load_workbook("content/交投桥梁基础数据-养护定检与年报编码匹配明细表2021.04.12（全部数据）.xlsx")

    sheet_check = book_check.active
    sheet_annual = book_annual["昆西-武昆176"]

    print(sheet_check)
    print(sheet_annual)

    find_loc(sheet_check[3], check_info, check_info_loc)
    find_loc(sheet_annual[3], annual_info, annual_info_loc)

    print(check_info_loc)
    print(annual_info_loc)

    # print("rowmax_ch:", sheet_check.max_row, "\trowmax_an:", sheet_annual.max_row)

    take_info(sheet_check, sheet_annual)
    book_check.save(filename = '桥梁动态数据-定期检查（云南云路工程检测有限公司-2020）_wukun_2.xlsx') 