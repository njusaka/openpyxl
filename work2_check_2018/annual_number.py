import openpyxl
import re
import datetime
if __name__ =="__main__":

    book_check = openpyxl.load_workbook("桥梁动态数据-定期检查（云南云路工程检测有限公司-2020）_final.xlsx")
    book_new = openpyxl.Workbook()
  
    sheet_check = book_check.active
    sheet_new = book_new.active

    route_name = ['安楚高速公路', '楚大高速公路', '楚广高速公路', '昆安高速公路', '昆明绕城公路西北段', '武昆高速公路', '武易高速公路', '永武高速公路']
    # print(sheet_check.max_row)
    bridge_name_old = ''
    bridge_num_old = ''
    row_new = 1
    for i in range(5, sheet_check.max_row+1):  #sheet_check.max_row+1
        bridge_route = sheet_check.cell(row=i, column=6).value   #bridge_route
        if bridge_route in route_name:
            # print('bridge_route:',bridge_route)
            bridge_name = sheet_check.cell(row=i, column=1).value   #bridge_name
            bridge_num = sheet_check.cell(row=i, column=7).value    #bridge_num
            bridge_an = sheet_check.cell(row=i, column=8).value     #bridge_an
            if bridge_name != bridge_name_old and bridge_num != bridge_num_old:
                print(row_new)
                sheet_new.cell(row=row_new, column=1).value = bridge_name
                sheet_new.cell(row=row_new, column=2).value = bridge_num
                sheet_new.cell(row=row_new, column=3).value = bridge_route
                sheet_new.cell(row=row_new, column=4).value = bridge_an
                row_new += 1
            bridge_name_old = bridge_name
            bridge_num_old = bridge_num
        else :
            continue 
    book_new.save(filename="annual_number.xlsx")




        

   