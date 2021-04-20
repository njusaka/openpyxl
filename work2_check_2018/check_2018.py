import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
import re
import getfile_more
source_sheet3_info = ['桥梁名称', '主跨结构', '桥长（m）', '路线名称', '桥梁编码', '本次检查日期' ]
source_sheet1_info = ['部件', '构件编号', '构件名称', '病害类型', '病害位置', '病害描述', '病害标度']
source_sheet3_info_loc = ['C4', 'G3', 'G4', 'C5', 'C3', 'L5']
source_sheet1_info_loc_col = [2, 3, 4, 5, 6, 8, 15]
info_original_1 = []
info_original_2 = []

b_up = '上行线'
b_down = '下行线'
bridge_a_special =[b_up, b_down, '人行天桥', '跨线桥', '线外桥', '上跨桥', '左幅', '右幅']

row_sheet1_record = 8
input_sheet_record = 5

class bridge:
    def __init__(self):
        self.name = None
        self.a = None
        self.bridge_type = "[梁式桥]"
        self.span = None
        self.bridge_long = None
        self.route_name = None
        self.num_stack = None
        self.annual = None
        self.check_date = None
        self.unit = "云南云路工程检测有限公司"
        self.stack_mark = None
        self.number = None

    def set_name_and_a(self):
        bridge_name_original = info_original_1[0]
        bridge_name_original = bridge_name_original.rstrip('桥')
        bridge_name_original = bridge_name_original.lstrip('匝')
        stack_m = re.sub(u'[\u4e00-\u9fa5]', '', bridge_name_original)
        bridge_n = re.sub(u'[^\u4e00-\u9fa5]', '', bridge_name_original)
        if '+' in stack_m:
            self.stack_mark = stack_m
            if bridge_n in bridge_a_special:
                self.name = self.stack_mark + '（' + bridge_n +'）'  #上下行之类加中文括号
                self.a = bridge_n
            else:
                self.name = self.stack_mark + bridge_n #没有桥幅的不需要
        elif b_up in bridge_name_original:
            t_p = re.sub(b_up, '', bridge_name_original)
            self.name = t_p + '（' + b_up +'）' 
            self.a = b_up
        elif b_down in bridge_name_original:
            t_p = re.sub(b_down, '', bridge_name_original)
            self.name = t_p + '（' + b_down +'）' 
            self.a = b_down
        #桥幅
        if self.a != None:
            self.a = self.a.rstrip('线')
        else:
            self.a = None
        print('bridge_name:', self.name)

    def set_span(self):  #中文去掉搬过去
        span_original = info_original_1[1]
        self.span = re.sub(u'[\u4e00-\u9fa5]', '', span_original)
        print('span:', self.span)

    def set_bridgelong(self):    #桥长
        self.bridge_long = info_original_1[2]
        print('bridge_long', self.bridge_long)

    def set_route_name(self):
        route_name_original = info_original_1[3]
        if route_name_original == '昆明～安宁高速公路':
            self.route_name = '昆安高速公路'
        elif route_name_original == 'G56杭瑞高速公路安宁至楚雄段':
            self.route_name = '安楚高速公路'
        elif route_name_original == '楚雄～广通高速公路':
            self.route_name = '楚广高速公路'
        elif route_name_original == '楚雄～大理高速公路':
            self.route_name = '楚大高速公路'
        elif route_name_original == '永仁～武定高速公路':
            self.route_name = '永武高速公路'
        elif route_name_original == '武定～昆明高速公路':
            self.route_name = '武昆高速公路'
        print('route_name', self.route_name)

    def set_num_stack(self):
        self.number = info_original_1[4]
        if self.stack_mark != None:     
            self.num_stack = self.number + '/' + self.stack_mark
        else :
            self.num_stack = self.number
        print('stack_mark', self.num_stack)

    def set_checkdate(self):
        self.check_date = info_original_1[5]
        print('check_date', self.check_date)

    def take_set_info_annual(self):
        for row_an in range(1, sheet_annual.max_row+1):
            val_an_num = str(sheet_annual.cell(row=row_an, column=2).value)
            val_an_num = val_an_num.replace(' ', '')
            val_an_num = re.sub(u'[\u4e00-\u9fa5]', '', val_an_num)
            if  self.num_stack == val_an_num  or val_an_num in self.num_stack:
                self.annual = sheet_annual.cell(row=row_an, column=4).value
                break
            else :
                self.annual = None
        print('annual', self.annual)

    def update_bridge(self):
        self.set_name_and_a()
        self.set_span()
        self.set_bridgelong()
        self.set_route_name()
        self.set_num_stack()
        self.set_checkdate()
        self.take_set_info_annual()

class disease:
    def __init__(self):
        self.ass = None
        self.component_num = None
        self.disease_type = None
        self.disease_describe = None
        self.disease_scale = None
    
    def set_desribe(self):
        d_loc  = str(info_original_2[4])
        if self.component_num in d_loc:
            disease_loc = re.sub(self.component_num, '', d_loc)
        else:
            disease_loc = d_loc
        disease_describe_original = str(info_original_2[5])
        self.disease_describe = self.component_num + self.disease_type + '，' + disease_loc + disease_describe_original + '。'

    def update_disease(self):
        self.ass = str(info_original_2[0])
        self.component_num = str(info_original_2[1]) + str(info_original_2[2])
        self.disease_type = str(info_original_2[3])
        self.disease_scale = str(info_original_2[6])
        self.set_desribe()

    def test(self):
        self.set_desribe()
        print("self.ass:", self.ass)
        print("self.component_num:", self.component_num)
        print("self.disease_type:", self.disease_type)
        print("self.disease_describe:", self.disease_describe)
        print("self.disease_scale:", self.disease_scale)

def take_info_bridge():
    info_original_1.clear()
    for i in range(0, len(source_sheet3_info_loc)):
        val = sheet_doc_3[source_sheet3_info_loc[i]].value
        print("val",val)
        if type(val) == str:
            val = val.replace('\n', '')

        if val == None and '2018年桥梁记录表（昆西）/武昆（176座）' in open_contentfile:
            loc_col = re.sub(r'\d', '', source_sheet3_info_loc[i])
            loc_col = column_index_from_string(loc_col)
            loc_row = re.sub(r'[A-Z]', '', source_sheet3_info_loc[i])
            loc_row = int(loc_row) + 1
            val = sheet_doc_3.cell(row=loc_row, column=loc_col).value

        if val == None and open_contentfile == '2018年桥梁记录表（昆西）/永武（433座）/K2611+051上行线（波纹管外露）.xlsx':
            val = '4×20m+4×19m+18.5m+24m+18.5m现浇箱梁'
        if type(val) == str:
            val = val.replace(' ', '')
        info_original_1.append(val)
    return info_original_1

def take_info_original_2(row):
    info_original_2.clear()
    for i in range(0, len(source_sheet1_info_loc_col)):
        val = sheet_doc_1.cell(row=row, column=source_sheet1_info_loc_col[i]).value
        if type(val) == str:
            val = val.replace(' ', '')
            val = val.replace('\n', '')
        info_original_2.append(val)
    return info_original_2

def update_info_original_2(row):
    global info_original_2 #改内部值需要
    temp = []
    for i in range(0, len(source_sheet1_info_loc_col)):
        val = sheet_doc_1.cell(row=row, column=source_sheet1_info_loc_col[i]).value
        if type(val) == str:
            val = val.replace(' ', '')
            val = val.replace('\n', '')
        temp.append(val)

    # print("temp:", temp)
    if temp[1] == None or temp[1] == '／' or temp[1] == '/':         #编号那行
        # print("此行空")
        return 2

    if temp[0] != None:
        info_original_2 = temp
        # print("info_original_2:", info_original_2)
    else:
        t_val = info_original_2[0]
        scale_old = info_original_2[6]
        info_original_2 = temp
        info_original_2[0] = t_val
        if temp[6] == None:
            info_original_2[6] = scale_old

    return info_original_2

#一开始从第5行输入  一行一行输入
def input_check(bridge, disease):
    bridge_val= [item_br for item_br in bridge.__dict__.items()] 
    disease_val= [item_di for item_di in disease.__dict__.items()]
    for col_check in range(1, 15+1):  #A-O
        if col_check <= 10:
            sheet_check.cell(row=input_sheet_record, column=col_check).value = bridge_val[col_check-1][1]
        else:
            sheet_check.cell(row=input_sheet_record, column=col_check).value = disease_val[col_check-10-1][1]
    # book_check.save("check.xlsx")

def input_1_sheet():
    global row_sheet1_record
    global input_sheet_record
    global br_info, di_info
    take_info_bridge()
    test_blank_num = 0
    while update_info_original_2(row_sheet1_record) == 2:   
        test_blank_num += 1 
        row_sheet1_record += 1
        if test_blank_num > (sheet_doc_3.max_row-7):
            return 2
    row_sheet1_record += 1

    print(info_original_1)
    br_info.update_bridge()
    di_info.update_disease()
    input_check(br_info, di_info)
    input_sheet_record += 1
    
    # print(info_original_2)
    for item_1_sheet in range(row_sheet1_record, sheet_doc_1.max_row):
        if update_info_original_2(item_1_sheet) != 2:
            di_info.update_disease()
            input_check(br_info, di_info)
            input_sheet_record += 1
            # print(info_original_2)
    return 0

def input_more_sheet():
    global row_sheet1_record
    global book_doc
    global sheet_doc_1, sheet_doc_3
    global open_contentfile
    for i_dir in range(0, 4):        #0-4
        getfile_more.choose_content(i_dir)
        print(getfile_more.dir_file)
        for i in range(0, len(getfile_more.file_ass)):  #0
            if getfile_more.dir_file == getfile_more.dir_1:
                open_contentfile = '2018年桥梁记录表（昆西）/昆安（31座）/昆安每座桥记录表汇总/'+getfile_more.file_ass[i]
            elif getfile_more.dir_file == getfile_more.dir_2:
                open_contentfile = '2018年桥梁记录表（昆西）/楚广（40座）/'+getfile_more.file_ass[i]
            elif getfile_more.dir_file == getfile_more.dir_3:
                open_contentfile = '2018年桥梁记录表（昆西）/武昆（176座）/'+getfile_more.file_ass[i]
            elif getfile_more.dir_file == getfile_more.dir_4:
                open_contentfile = '2018年桥梁记录表（昆西）/永武（433座）/'+getfile_more.file_ass[i]
            print(i, open_contentfile)
            book_doc = openpyxl.load_workbook(open_contentfile)
            row_sheet1_record = 8
            if len(book_doc.sheetnames) == 3:  #3个表的情况
                sheet_doc_1 = book_doc[book_doc.sheetnames[0]]
                sheet_doc_3 = book_doc[book_doc.sheetnames[2]]
                if 'Sheet' in str(sheet_doc_3.title):
                    sheet_doc_3 = book_doc[book_doc.sheetnames[1]]
            elif len(book_doc.sheetnames) == 2:  #2个表的情况
                sheet_doc_1 = book_doc[book_doc.sheetnames[0]]
                sheet_doc_3 = book_doc[book_doc.sheetnames[1]]
            if input_1_sheet() == 2:
                continue

if __name__ =="__main__":

    book_check = openpyxl.load_workbook("桥梁动态数据-定期检查（云南云路工程检测有限公司-2018）.xlsx")
    book_annual = openpyxl.load_workbook("annual_number.xlsx")
    sheet_check = book_check.active
    sheet_annual = book_annual.active


    br_info = bridge()
    di_info = disease()

    input_more_sheet()

    book_check.save("check_4.xlsx")